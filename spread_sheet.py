from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, cell, column_index_from_string
import time
import os
import sqlite3


connect = sqlite3.connect('face-database.db')
c = connect.cursor()
currentDate = time.strftime("%d_%m_%y")

if(os.path.exists('./reports.xlsx')):
    wb = load_workbook(filename = "reports.xlsx") # as workbook contains multiple sheets 
    sheet = wb['IT2016']
    #print(type(currentDate))
    if sheet.cell(row=1,column=3).value != currentDate:   #when u have already created the worksheet then we have to update the attendance 
        sheet['C1'].value=currentDate
    #saving the file
    wb.save(filename = "reports.xlsx")

else:   # here you are appending all the database to spreadsheet 
    wb = Workbook()
    dest_filename = 'reports.xlsx'
    c.execute("select * from students order by Roll ASC") # that will all rows sorted by roll number 
    #creating worksheet and giving names to column
    ws1 = wb.active
    ws1.title = "IT2016"  # this would be the title of the worksheet 
    ws1.append(('Roll Number','Name',currentDate)) # name of the columns 
    #ws1.append(('','',''))
    #entering students information from the database
    while True:
        row = c.fetchone()
        if row == None:
            break 
        else:
            ws1.append((row[2],row[1])) #roll number #name 
    #saving the file
    wb.save(filename = dest_filename)






    